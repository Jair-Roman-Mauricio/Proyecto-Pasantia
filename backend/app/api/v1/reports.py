import io
from datetime import datetime
from decimal import Decimal

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from app.dependencies import check_permission
from app.subest_client import get_subest_client

router = APIRouter(prefix="/reports", tags=["Reports"])


def _get_demand_data(client, start_date=None, end_date=None):
    stations = client.table("stations").select("*").order("order_index").execute().data

    if not start_date and not end_date:
        return [
            {
                "station_id": s["id"],
                "station_name": s["name"],
                "station_code": s["code"],
                "transformer_capacity_kw": float(s["transformer_capacity_kw"]),
                "max_demand_kw": float(s["max_demand_kw"]),
                "available_power_kw": float(s["available_power_kw"]),
                "status": s["status"],
            }
            for s in stations
        ]

    data = []
    for s in stations:
        bars = client.table("bars").select("id").eq("station_id", s["id"]).execute().data
        bar_ids = [b["id"] for b in bars]
        total_md = Decimal("0")

        if bar_ids:
            circ_q = client.table("circuits").select("id,md_kw").in_("bar_id", bar_ids).neq("status", "inactive")
            if start_date:
                circ_q = circ_q.gte("created_at", start_date.isoformat())
            if end_date:
                circ_q = circ_q.lte("created_at", end_date.isoformat())
            circuits = circ_q.execute().data
            total_md += sum(Decimal(str(c["md_kw"])) for c in circuits)

            circuit_ids = [c["id"] for c in circuits]
            if circuit_ids:
                sub_q = client.table("sub_circuits").select("md_kw").in_("circuit_id", circuit_ids).eq("status", "operative_normal")
                if start_date:
                    sub_q = sub_q.gte("created_at", start_date.isoformat())
                if end_date:
                    sub_q = sub_q.lte("created_at", end_date.isoformat())
                subs = sub_q.execute().data
                total_md += sum(Decimal(str(s2["md_kw"])) for s2 in subs)

        capacity = float(s["transformer_capacity_kw"])
        demand = float(total_md)
        data.append({
            "station_id": s["id"],
            "station_name": s["name"],
            "station_code": s["code"],
            "transformer_capacity_kw": capacity,
            "max_demand_kw": demand,
            "available_power_kw": capacity - demand,
            "status": s["status"],
        })
    return data


@router.get("/demand-evolution")
def get_demand_evolution(start_date: datetime | None = None, end_date: datetime | None = None, _=Depends(check_permission("view_reports"))):
    client = get_subest_client()
    return _get_demand_data(client, start_date, end_date)


@router.get("/requests-per-station")
def get_requests_per_station(start_date: datetime | None = None, end_date: datetime | None = None, _=Depends(check_permission("view_reports"))):
    client = get_subest_client()
    req_q = client.table("requests").select("station_id,status")
    if start_date:
        req_q = req_q.gte("created_at", start_date.isoformat())
    if end_date:
        req_q = req_q.lte("created_at", end_date.isoformat())
    requests = req_q.execute().data

    stations = {s["id"]: s["name"] for s in client.table("stations").select("id,name").execute().data}
    data: dict = {}
    for r in requests:
        sid = r["station_id"]
        name = stations.get(sid, str(sid))
        if name not in data:
            data[name] = {"station_id": sid, "station_name": name, "pending": 0, "approved": 0, "rejected": 0}
        data[name][r["status"]] = data[name].get(r["status"], 0) + 1
    return list(data.values())


@router.get("/export/excel")
def export_reports_excel(start_date: datetime | None = None, end_date: datetime | None = None, _=Depends(check_permission("view_reports"))):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font

    client = get_subest_client()
    wb = Workbook()
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws1 = wb.active
    ws1.title = "Demanda Electrica"
    ws1.append(["Estacion", "Codigo", "Capacidad (kW)", "Demanda Actual (kW)", "Energia Disponible (kW)", "Estado"])
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font

    status_labels = {"green": "Energia suficiente", "yellow": "Menos del 20% disponible", "red": "Debe energia"}
    for row in _get_demand_data(client, start_date, end_date):
        ws1.append([row["station_name"], row["station_code"], row["transformer_capacity_kw"],
                    row["max_demand_kw"], row["available_power_kw"], status_labels.get(row["status"], row["status"])])
    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].width = 24

    ws2 = wb.create_sheet("Solicitudes por Estacion")
    ws2.append(["Estacion", "Pendientes", "Aprobadas", "Rechazadas", "Total"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in get_requests_per_station(start_date, end_date, _):
        total = row["pending"] + row["approved"] + row["rejected"]
        ws2.append([row["station_name"], row["pending"], row["approved"], row["rejected"], total])
    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = "reportes.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})
